# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# License: MIT. See LICENSE
import datetime
import re
from io import BytesIO
from typing import Any

import openpyxl
import xlrd
from openpyxl import load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.child import INVALID_TITLE_REGEX

import frappe
from frappe.core.utils import html2text
from frappe.utils.html_utils import unescape_html

ILLEGAL_CHARACTERS_RE = re.compile(
	r"[\000-\010]|[\013-\014]|[\016-\037]|\uFEFF|\uFFFE|\uFFFF|[\uD800-\uDFFF]"
)


def get_excel_date_format():
	date_format = frappe.get_system_settings("date_format")
	time_format = frappe.get_system_settings("time_format")

	# Excel-compatible format
	date_format = date_format.replace("mm", "MM")

	return date_format, time_format


# return xlsx file object
def make_xlsx(
	data: list[list[Any]],
	sheet_name: str,
	wb: openpyxl.Workbook | None = None,
	column_widths: list[int] | None = None,
	header_index: int = 0,
	has_filters: bool = False,
) -> BytesIO:
	"""
	Create an Excel file with the given data and formatting options.

	Args:
		data: List of rows, where each row is a list of cell values
		sheet_name: Name of the Excel sheet
		wb: Existing workbook to add sheet to. If None, creates new workbook
		column_widths: List of column widths in Excel units. If None, auto-sized
		header_index: Row index (0-based) that should be formatted as header making it bold
		has_filters: If True, applies bold formatting to the first column of filter rows

	Returns:
		BytesIO: object containing the Excel file data
	"""
	column_widths = column_widths or []
	if wb is None:
		wb = openpyxl.Workbook(write_only=True)

	sheet_name_sanitized = INVALID_TITLE_REGEX.sub(" ", sheet_name)
	ws = wb.create_sheet(sheet_name_sanitized, 0)

	for i, column_width in enumerate(column_widths):
		if column_width:
			ws.column_dimensions[get_column_letter(i + 1)].width = column_width

	date_format, time_format = get_excel_date_format()
	bold_font = Font(name="Calibri", bold=True)

	for row_idx, row in enumerate(data):
		clean_row = []
		is_header_row = row_idx == header_index
		is_filter_row = has_filters and row_idx < header_index

		for col_idx, item in enumerate(row):
			if isinstance(item, str) and (sheet_name not in ["Data Import Template", "Data Export"]):
				value = handle_html(item)
			else:
				value = item

			if isinstance(item, str) and next(ILLEGAL_CHARACTERS_RE.finditer(value), None):
				# Remove illegal characters from the string
				value = ILLEGAL_CHARACTERS_RE.sub("", value)

			cell = WriteOnlyCell(ws, value=value)

			if isinstance(value, datetime.date | datetime.datetime):
				number_format = date_format
				if isinstance(value, datetime.datetime):
					number_format = f"{date_format} {time_format}"
				cell.number_format = number_format

			# Apply bold font for header row or first column of filter rows
			if is_header_row or (is_filter_row and col_idx == 0):
				cell.font = bold_font

			clean_row.append(cell)

		ws.append(clean_row)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file


### Utilities ###
def handle_html(data: str) -> str:
	# return if no html tags found
	if "<" not in data or ">" not in data:
		return data

	h = unescape_html(data or "")

	try:
		value = html2text(h, strip_links=True, wrap=False)
	except Exception:
		# unable to parse html, send it raw
		return data

	return value.replace("  \n", ", ").replace("\n", " ").replace("# ", ", ")


def read_xlsx_file_from_attached_file(file_url=None, fcontent=None, filepath=None):
	if file_url:
		_file = frappe.get_doc("File", {"file_url": file_url})
		filename = _file.get_full_path()
	elif fcontent:
		filename = BytesIO(fcontent)
	elif filepath:
		filename = filepath
	else:
		return

	rows = []
	wb1 = load_workbook(filename=filename, data_only=True)
	ws1 = wb1.active
	for row in ws1.iter_rows():
		rows.append([cell.value for cell in row])
	return rows


def read_xls_file_from_attached_file(content):
	book = xlrd.open_workbook(file_contents=content)
	sheets = book.sheets()
	sheet = sheets[0]
	return [sheet.row_values(i) for i in range(sheet.nrows)]


def build_xlsx_response(data, filename):
	from frappe.desk.utils import provide_binary_file

	provide_binary_file(filename, "xlsx", make_xlsx(data, filename).getvalue())
